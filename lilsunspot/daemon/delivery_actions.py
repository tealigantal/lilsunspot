from __future__ import annotations

import csv
import io
import re
import secrets
import zipfile
from base64 import b64decode
from contextlib import contextmanager
from contextvars import ContextVar
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator
from xml.sax.saxutils import escape as xml_escape

from . import conversations
from .attachments import DESKTOP_UPLOAD_MAX_BYTES, is_safe_stored_attachment
from .config_paths import RuntimePaths


_CURRENT_TURN: ContextVar["DeliveryTurnContext | None"] = ContextVar("lilsunspot_delivery_turn", default=None)
_ATTACHMENT_ID_RE = re.compile(r"^att_[A-Za-z0-9_-]+$")
_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
_SAFE_FILE_NAME_RE = re.compile(r"[^A-Za-z0-9._\-\u4e00-\u9fff ]+")
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_PDF_MIME = "application/pdf"
_STRICT_FORMAT_MIME_BY_EXT = {
    ".xlsx": _XLSX_MIME,
    ".docx": _DOCX_MIME,
    ".pdf": _PDF_MIME,
}
_TEXT_MIME_BY_EXT = {
    ".csv": "text/csv",
    ".txt": "text/plain",
    ".md": "text/markdown",
    ".json": "application/json",
    ".yaml": "application/yaml",
    ".yml": "application/yaml",
    ".log": "text/plain",
}


@dataclass
class DeliveryTurnContext:
    conversation_id: str
    source: str
    route: dict[str, str] | None
    paths: RuntimePaths
    deliverable_dir: Path
    actions: list[dict[str, Any]] = field(default_factory=list)
    _action_by_attachment: dict[str, dict[str, Any]] = field(default_factory=dict)
    _action_by_path: dict[str, dict[str, Any]] = field(default_factory=dict)

    @property
    def target(self) -> str:
        return "weixin_current_route" if self.route else "desktop_current_conversation"

    def add_action(self, action: dict[str, Any]) -> dict[str, Any]:
        attachment_id = str(action.get("attachment_id") or "")
        if action.get("ok") and attachment_id:
            existing = self._action_by_attachment.get(attachment_id)
            if existing is not None:
                return _public_action({**existing, "duplicate": True})
            self._action_by_attachment[attachment_id] = action
        self.actions.append(action)
        return _public_action(action)

    def add_file_action(self, action: dict[str, Any]) -> dict[str, Any]:
        safe_path = str(action.get("safe_path") or "")
        if action.get("ok") and safe_path:
            key = str(Path(safe_path).resolve(strict=False)).lower()
            existing = self._action_by_path.get(key)
            if existing is not None:
                return _public_action({**existing, "duplicate": True})
            self._action_by_path[key] = action
        self.actions.append(action)
        return _public_action(action)

    def actions_for_result(self) -> list[dict[str, Any]]:
        return [dict(action) for action in self.actions]


def current_delivery_turn() -> DeliveryTurnContext | None:
    return _CURRENT_TURN.get()


@contextmanager
def delivery_turn_context(
    *,
    conversation_id: str,
    source: str,
    route: dict[str, str] | None,
    paths: RuntimePaths,
    deliverable_dir: Path | None = None,
) -> Iterator[DeliveryTurnContext]:
    resolved_deliverable_dir = deliverable_dir or deliverable_dir_for_turn(paths, conversation_id)
    resolved_deliverable_dir.mkdir(parents=True, exist_ok=True)
    context = DeliveryTurnContext(
        conversation_id=conversation_id,
        source=source,
        route=route,
        paths=paths,
        deliverable_dir=resolved_deliverable_dir,
    )
    token = _CURRENT_TURN.set(context)
    try:
        yield context
    finally:
        _CURRENT_TURN.reset(token)


def return_attachment_action(attachment_id: str, caption: str = "") -> dict[str, Any]:
    context = current_delivery_turn()
    if context is None:
        return _failed_action(attachment_id, "no_active_turn")

    attachment_id = str(attachment_id or "").strip()
    if not _ATTACHMENT_ID_RE.match(attachment_id):
        return context.add_action(_failed_action(attachment_id, "invalid_attachment_id", target=context.target))

    attachment = conversations.get_attachment(attachment_id, include_safe_path=True, paths=context.paths)
    if not attachment:
        return context.add_action(_failed_action(attachment_id, "missing_attachment", target=context.target))

    if str(attachment.get("conversation_id") or "") != context.conversation_id:
        return context.add_action(_failed_action(attachment_id, "cross_conversation", target=context.target))

    try:
        safe_path = is_safe_stored_attachment(str(attachment.get("safe_path") or ""), context.paths)
    except Exception:
        return context.add_action(_failed_action(attachment_id, "unsafe_path", target=context.target))

    action = {
        "ok": True,
        "action_id": f"da_{secrets.token_hex(8)}",
        "attachment_id": attachment_id,
        "safe_path": str(safe_path),
        "file_name": str(attachment.get("file_name") or Path(safe_path).name),
        "mime_type": str(attachment.get("mime_type") or "application/octet-stream"),
        "media_kind": _media_kind(attachment, safe_path),
        "caption": _clean_caption(caption),
        "target": context.target,
        "status": "pending",
    }
    return context.add_action(action)


def deliver_file_action(path: str, caption: str = "") -> dict[str, Any]:
    context = current_delivery_turn()
    if context is None:
        return _failed_action("", "no_active_turn")

    try:
        safe_path = _resolve_deliverable_path(path, context)
    except ValueError as exc:
        return context.add_file_action(_failed_file_action(path, str(exc), target=context.target))
    if not safe_path.is_file():
        return context.add_file_action(_failed_file_action(path, "missing_file", target=context.target))
    try:
        size_bytes = safe_path.stat().st_size
    except OSError:
        return context.add_file_action(_failed_file_action(path, "missing_file", target=context.target))
    if size_bytes <= 0:
        return context.add_file_action(_failed_file_action(path, "empty_file", target=context.target))
    if size_bytes > DESKTOP_UPLOAD_MAX_BYTES:
        return context.add_file_action(_failed_file_action(path, "file_too_large", target=context.target))
    format_reason = validate_deliverable_file_for_delivery(safe_path)
    if format_reason:
        return context.add_file_action(_failed_file_action(path, format_reason, target=context.target))
    return context.add_file_action(_file_action_for_path(safe_path, context, caption=caption))


def create_deliverable_file_action(
    *,
    file_name: str,
    content_text: str | None = None,
    content_base64: str | None = None,
    mime_type: str = "",
    caption: str = "",
) -> dict[str, Any]:
    context = current_delivery_turn()
    if context is None:
        return _failed_action("", "no_active_turn")

    safe_name = _safe_file_name(file_name)
    if not safe_name:
        return context.add_file_action(_failed_action("", "invalid_file_name", target=context.target))
    if content_text is not None and content_base64 is not None:
        return context.add_file_action(_failed_action("", "ambiguous_file_content", target=context.target))
    if content_text is None and content_base64 is None:
        return context.add_file_action(_failed_action("", "missing_file_content", target=context.target))

    data, inferred_mime, content_reason = _materialize_file_content(
        safe_name,
        content_text=content_text,
        content_base64=content_base64,
    )
    if content_reason:
        return context.add_file_action(_failed_action("", content_reason, target=context.target))
    if not data:
        return context.add_file_action(_failed_action("", "empty_file", target=context.target))
    if len(data) > DESKTOP_UPLOAD_MAX_BYTES:
        return context.add_file_action(_failed_action("", "file_too_large", target=context.target))

    target_path = _unique_path(context.deliverable_dir / safe_name)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        target_path.write_bytes(data)
    except OSError:
        return context.add_file_action(_failed_action("", "file_write_failed", target=context.target))
    format_reason = validate_deliverable_file_for_delivery(target_path)
    if format_reason:
        _remove_file_quietly(target_path)
        return context.add_file_action(_failed_file_action(target_path.name, format_reason, target=context.target))
    return context.add_file_action(
        _file_action_for_path(target_path, context, caption=caption, mime_type=mime_type or inferred_mime)
    )


def _failed_action(attachment_id: str, reason_code: str, *, target: str = "") -> dict[str, Any]:
    return {
        "ok": False,
        "action_id": f"da_{secrets.token_hex(8)}",
        "attachment_id": str(attachment_id or "").strip(),
        "file_name": "",
        "mime_type": "",
        "media_kind": "",
        "caption": "",
        "target": target,
        "status": "rejected",
        "reason_code": reason_code,
    }


def _failed_file_action(path: str, reason_code: str, *, target: str = "") -> dict[str, Any]:
    action = _failed_action("", reason_code, target=target)
    action["file_name"] = Path(str(path or "")).name
    return action


def _file_action_for_path(
    safe_path: Path,
    context: DeliveryTurnContext,
    *,
    caption: str = "",
    mime_type: str = "",
) -> dict[str, Any]:
    mime = _preferred_mime_type(safe_path, mime_type)
    media_kind = "image" if mime.lower().startswith("image/") or safe_path.suffix.lower() in _IMAGE_EXTS else "document"
    return {
        "ok": True,
        "action_id": f"da_{secrets.token_hex(8)}",
        "attachment_id": "",
        "safe_path": str(safe_path),
        "file_name": safe_path.name,
        "mime_type": mime,
        "media_kind": media_kind,
        "caption": _clean_caption(caption),
        "target": context.target,
        "status": "pending",
    }


def _media_kind(attachment: dict[str, Any], safe_path: Path) -> str:
    mime_type = str(attachment.get("mime_type") or "").lower()
    if mime_type.startswith("image/") or safe_path.suffix.lower() in _IMAGE_EXTS:
        return "image"
    return "document"


def deliverable_dir_for_turn(paths: RuntimePaths, conversation_id: str, turn_id: str | None = None) -> Path:
    return (
        paths.hermes_home
        / "cache"
        / "documents"
        / _safe_path_part(conversation_id or "conversation")
        / _safe_path_part(turn_id or f"turn_{secrets.token_hex(6)}")
    )


def _safe_path_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip()).strip("._-")
    return cleaned[:80] or "conversation"


def _safe_file_name(value: str) -> str:
    name = Path(str(value or "").replace("\\", "/")).name.strip().strip(".")
    name = _SAFE_FILE_NAME_RE.sub("_", name)
    name = re.sub(r"\s+", " ", name).strip()
    if not name:
        return ""
    path = Path(name)
    stem = (path.stem or "file")[:120].strip() or "file"
    suffix = path.suffix[:20]
    return f"{stem}{suffix}" if suffix else stem


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(2, 1000):
        candidate = path.with_name(f"{stem}-{index}{suffix}")
        if not candidate.exists():
            return candidate
    return path.with_name(f"{stem}-{secrets.token_hex(4)}{suffix}")


def _resolve_deliverable_path(path: str, context: DeliveryTurnContext) -> Path:
    if not str(path or "").strip():
        raise ValueError("missing_file_path")
    candidate = Path(str(path)).expanduser()
    if not candidate.is_absolute():
        candidate = context.deliverable_dir / candidate
    try:
        resolved = candidate.resolve(strict=False)
        root = context.deliverable_dir.resolve(strict=False)
        resolved.relative_to(root)
    except Exception as exc:
        raise ValueError("unsafe_path") from exc
    return resolved


def _guess_mime_type(path: Path) -> str:
    import mimetypes

    guessed, _encoding = mimetypes.guess_type(str(path))
    return guessed or "application/octet-stream"


def _preferred_mime_type(path: Path, provided_mime: str = "") -> str:
    suffix = path.suffix.lower()
    if suffix in _STRICT_FORMAT_MIME_BY_EXT:
        return _STRICT_FORMAT_MIME_BY_EXT[suffix]
    if suffix in _TEXT_MIME_BY_EXT:
        return _TEXT_MIME_BY_EXT[suffix]
    return str(provided_mime or _guess_mime_type(path))


def _materialize_file_content(
    file_name: str,
    *,
    content_text: str | None,
    content_base64: str | None,
) -> tuple[bytes, str, str]:
    suffix = Path(file_name).suffix.lower()
    if content_text is not None:
        text = str(content_text)
        try:
            if suffix == ".xlsx":
                return _xlsx_bytes_from_text(text), _XLSX_MIME, ""
            if suffix == ".docx":
                return _docx_bytes_from_text(text), _DOCX_MIME, ""
        except Exception:
            return b"", "", "unsupported_generated_format"
        if suffix == ".pdf":
            return b"", "", "unsupported_generated_format"
        return text.encode("utf-8"), _TEXT_MIME_BY_EXT.get(suffix, ""), ""

    try:
        data = _decode_base64_payload(content_base64)
    except Exception:
        return b"", "", "invalid_base64"
    reason = _validate_strict_format_bytes(file_name, data)
    if reason:
        return b"", "", reason
    return data, _STRICT_FORMAT_MIME_BY_EXT.get(suffix, _TEXT_MIME_BY_EXT.get(suffix, "")), ""


def _rows_from_text(text: str) -> list[list[str]]:
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    if not lines:
        return [["内容"], [""]]

    markdown_rows: list[list[str]] = []
    for line in lines:
        if "|" not in line:
            continue
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if cells and all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) for cell in cells):
            continue
        if any(cells):
            markdown_rows.append(cells)
    if markdown_rows:
        return markdown_rows

    sample = "\n".join(lines[:50])
    delimiter = "\t" if "\t" in sample else "," if "," in sample else ""
    if delimiter:
        rows = list(csv.reader(io.StringIO("\n".join(lines)), delimiter=delimiter))
        return [[str(cell) for cell in row] for row in rows if any(str(cell).strip() for cell in row)]
    return [["内容"], *[[line] for line in lines]]


def _xlsx_bytes_from_text(text: str) -> bytes:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("openpyxl unavailable") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row in _rows_from_text(text):
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _docx_bytes_from_text(text: str) -> bytes:
    lines = str(text or "").splitlines() or [""]
    paragraph_xml = "".join(
        f'<w:p><w:r><w:t xml:space="preserve">{xml_escape(line)}</w:t></w:r></w:p>' for line in lines
    )
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{paragraph_xml}<w:sectPr/></w:body>"
        "</w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/>'
        "</Relationships>"
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as package:
        package.writestr("[Content_Types].xml", content_types)
        package.writestr("_rels/.rels", rels)
        package.writestr("word/document.xml", document_xml)
    return buffer.getvalue()


def validate_deliverable_file_for_delivery(path: str | Path) -> str:
    candidate = Path(path)
    try:
        if not candidate.is_file():
            return "missing_file"
        size_bytes = candidate.stat().st_size
    except OSError:
        return "missing_file"
    if size_bytes <= 0:
        return "empty_file"
    if size_bytes > DESKTOP_UPLOAD_MAX_BYTES:
        return "file_too_large"
    return _validate_strict_format_file(candidate)


def _validate_strict_format_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix not in _STRICT_FORMAT_MIME_BY_EXT:
        return ""
    try:
        data = path.read_bytes()
    except OSError:
        return "missing_file"
    return _validate_strict_format_bytes(path.name, data)


def _validate_strict_format_bytes(file_name: str, data: bytes) -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".xlsx":
        return _validate_xlsx_bytes(data)
    if suffix == ".docx":
        return _validate_docx_bytes(data)
    if suffix == ".pdf":
        return _validate_pdf_bytes(data)
    return ""


def _validate_xlsx_bytes(data: bytes) -> str:
    if not data.startswith(b"PK"):
        return "invalid_file_format"
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        workbook.close()
    except Exception:
        return "invalid_file_format"
    return ""


def _validate_docx_bytes(data: bytes) -> str:
    if not data.startswith(b"PK"):
        return "invalid_file_format"
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as package:
            names = set(package.namelist())
            if "[Content_Types].xml" not in names or "word/document.xml" not in names:
                return "invalid_file_format"
            package.read("word/document.xml")
    except (OSError, zipfile.BadZipFile, KeyError):
        return "invalid_file_format"
    return ""


def _validate_pdf_bytes(data: bytes) -> str:
    if not data.lstrip().startswith(b"%PDF-"):
        return "invalid_file_format"
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(data))
        if len(reader.pages) < 1:
            return "invalid_file_format"
    except ImportError:
        if b"%%EOF" not in data[-2048:]:
            return "invalid_file_format"
    except Exception:
        return "invalid_file_format"
    return ""


def _remove_file_quietly(path: Path) -> None:
    try:
        path.unlink()
    except OSError:
        pass


def _decode_base64_payload(value: str | None) -> bytes:
    raw = str(value or "").strip()
    if raw.lower().startswith("data:") and "," in raw:
        raw = raw.split(",", 1)[1].strip()
    return b64decode(raw, validate=True)


def _clean_caption(caption: str) -> str:
    value = " ".join(str(caption or "").split())
    if len(value) > 240:
        value = value[:240].rstrip()
    return value


def _public_action(action: dict[str, Any]) -> dict[str, Any]:
    allowed = {
        "ok",
        "action_id",
        "attachment_id",
        "file_name",
        "mime_type",
        "media_kind",
        "caption",
        "target",
        "status",
        "reason_code",
        "duplicate",
    }
    return {key: value for key, value in action.items() if key in allowed}
